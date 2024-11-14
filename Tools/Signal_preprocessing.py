import warnings
from openpyxl.workbook import Workbook
from scipy.interpolate import interp1d

warnings.filterwarnings("ignore")
import os
import pickle
import numpy as np
import pandas as pd
import scipy as sp
from scipy.signal import butter, filtfilt
from tqdm import tqdm
import xlrd
import xlwt
from scipy.signal import resample, decimate

from scipy.signal import resample

# Butterworth filter
def butterworth_offline_filter(data, func, cutoff1=None, cutoff2=None, fs=None, order=None):
    filter_df = pd.DataFrame()
    # Extract features for each sample according to the original index
    for index_name in tqdm(data.index):
        filter_data = func(data.loc[index_name, :], cutoff1=cutoff1, cutoff2=cutoff2, fs=fs, order=order)
        filter_data = pd.DataFrame(filter_data).T  # Transpose
        filter_data.index = [index_name]
        filter_df = pd.concat([filter_df, filter_data], axis=0)  # Concatenate rows
    return filter_df

'''
Assume a sampling frequency of 1000 Hz, the maximum frequency of the signal itself is 500 Hz. 
To filter out frequency components above 400 Hz, set the cutoff frequency to 400 Hz, so wn=2*400/1000=0.8. 
Thus, Wn=0.8.
fs: Sampling frequency
cutoff1: Filter out frequency components above cutoff1
'''
# Define low-pass filter function
def butter_lowpass_filter(data, cutoff, fs, order=4):
    nyquist = 0.5 * fs  # Nyquist frequency
    normal_cutoff = cutoff / nyquist  # Normalized cutoff frequency
    b, a = butter(order, normal_cutoff, btype='low', analog=False)  # Create filter
    y = data
    if len(data) > 15:
        y = filtfilt(b, a, data)  # Apply filter
    return y

# Define function: Segment each column of data for filtering (if there are gaps in timestamps, the data will be left empty)
def filter_column(column, cutoff, fs):
    filtered_col = column.copy()  # Copy the column
    isnan = column.isna()  # Locate NaN values
    valid_segments = []  # Store index ranges of valid data segments

    # Identify valid data segments that do not contain NaN
    valid_data = ~isnan
    start = None
    for i in range(len(column)):
        if valid_data[i] and start is None:
            start = i  # Record start of valid data segment
        elif not valid_data[i] and start is not None:
            valid_segments.append((start, i))  # Record end of valid data segment
            start = None
    if start is not None:
        valid_segments.append((start, len(column)))  # Process the last valid data segment

    # Apply filter to each valid data segment
    for start, end in valid_segments:
        filtered_col[start:end] = butter_lowpass_filter(column[start:end], cutoff, fs)

    return filtered_col

'''
Assume a sampling frequency of 1000 Hz, the maximum frequency of the signal itself is 500 Hz. 
To filter out frequency components below 100 Hz, set the cutoff frequency to 100 Hz, so wn=2*100/1000=0.2. 
Thus, Wn=0.2.
'''
def butter_highpass_filter(data, cutoff1=50, cutoff2=None, fs=128, order=5):
    '''
    param:data: Original signal
    return:filtered_data: Filtered signal
    '''
    wn = 2 * cutoff1 / fs  # High-pass boundary
    b, a = butter(N=order, Wn=wn, btype='highpass')
    filtered_data = filtfilt(b, a, data)
    return filtered_data

'''
Assume a sampling frequency of 1000 Hz, the maximum frequency of the signal itself is 500 Hz. 
To filter out frequency components below 100 Hz and above 400 Hz, set the cutoff frequencies to 100 Hz and 400 Hz, 
so wn1=2*100/1000=0.2, Wn1=0.2; wn2=2*400/1000=0.8, Wn2=0.8. Wn=[0.02,0.8]
'''
def butter_bandpass_filter(data, cutoff1=1, cutoff2=50, fs=128, order=5):
    '''
    param:data: Original signal
    return:filtered_data: Filtered signal
    '''
    wn1 = 2 * cutoff1 / fs  # Lower cutoff frequency
    wn2 = 2 * cutoff2 / fs  # Upper cutoff frequency
    wn = [wn1, wn2]  # Band-pass range

    b, a = butter(N=order, Wn=wn, btype='bandpass')
    filtered_data = filtfilt(b, a, data)
    return filtered_data

'''
Assume a sampling frequency of 1000 Hz, the maximum frequency of the signal itself is 500 Hz. 
To filter out frequency components below 100 Hz and above 400 Hz, set the cutoff frequencies to 100 Hz and 400 Hz, 
so wn1=2*100/1000=0.2, Wn1=0.2; wn2=2*400/1000=0.8, Wn2=0.8. Wn=[0.02,0.8]. 
This is similar to band-pass, but the band-pass retains the middle, while the band-stop removes it.
'''
def butter_bandstop_filter(data, cutoff1=1, cutoff2=50, fs=128, order=5):
    '''
    param:data: Original signal
    return:filtered_data: Filtered signal
    '''
    wn1 = 2 * cutoff1 / fs  # Lower cutoff frequency
    wn2 = 2 * cutoff2 / fs  # Upper cutoff frequency
    wn = [wn1, wn2]  # Band-stop range

    b, a = butter(N=order, Wn=wn, btype='bandstop')
    filtered_data = filtfilt(b, a, data)
    return filtered_data

# Downsampling
'''
First, calculate the decimation_factor, which is the ratio of the original sampling rate to the target sampling rate. 
Then use signal.butter to design a low-pass filter with a cutoff frequency set to half the target frequency. 
Apply the filter to remove aliasing errors, and finally downsample the signal with Python's slicing syntax.
data: Data to be downsampled
fs: Original sampling frequency
fs_new: New sampling frequency
'''
def down_sampling1(data, fs, fs_new):
    from scipy import signal

    # Calculate downsampling and upsampling factors
    downsample_factor = int(np.round(fs / fs_new))
    upsample_factor = int(np.round(fs_new / fs))

    # FIR filter
    nyquist = 0.5 * fs
    cutoff = 0.45 * nyquist
    b = signal.firwin(101, cutoff / nyquist)
    x_filtered = signal.lfilter(b, 1, data)

    # Resample using upsampling and downsampling
    x_upsampled = signal.resample_poly(x_filtered, upsample_factor, 1)
    x_downsampled = signal.resample_poly(x_upsampled, 1, downsample_factor)
    return x_downsampled

def down_sampling2(data, fs, fs_new):
    # Set downsampling factor
    data_new = []
    x = int(fs/(fs-fs_new))+1
    for i in range(len(data)):
        if not i % x == 0:
            data_new.append(data[i])
    return data_new

def down_sampling3(data, original_sampling_rate, target_sampling_rate):
    # Downsampling from 120Hz to 100Hz
    num_samples = int(len(data) * (target_sampling_rate / original_sampling_rate))
    downsampled_data = resample(data, num_samples)
    return downsampled_data

def down_sampling4(data, original_sampling_rate, target_sampling_rate):
    result = pd.DataFrame()
    data = data.iloc[:, :16]
    for column in data.columns:
        series = data[column].values
        original_length = len(series)

        # Calculate downsampled length
        downsampled_length = int(original_length * (target_sampling_rate / original_sampling_rate))
        result_series = np.full(downsampled_length, np.nan)

        # Locate valid data indices
        valid_indices = np.where(~np.isnan(series))[0]

        # If no valid data, return all NaN
        if len(valid_indices) == 0:
            result[column] = result_series
            continue

        # Segment valid data for processing
        segments = []
        starts = []
        start = 0
        for idx in range(original_length):
            if np.isnan(series[idx]):
                if start < idx:
                    segments.append(series[start:idx])
                    starts.append(start)  # Record start of segment
                start = idx + 1
        if start < original_length:
            segments.append(series[start:])
            starts.append(start)

        # Downsample each valid segment
        for segment, segment_start in zip(segments, starts):
            segment_length = len(segment)
            if segment_length > 0:
                segment_downsampled_length = int(segment_length * (target_sampling_rate / original_sampling_rate))
                if segment_downsampled_length > 0:
                    downsampled_segment = resample(segment, segment_downsampled_length)
                    segment_start_index = int(segment_start * (target_sampling_rate / original_sampling_rate))
                    segment_end_index = segment_start_index + segment_downsampled_length
                    result_series[segment_start_index:segment_end_index] = downsampled_segment

        # Add result to DataFrame
        result[column] = result_series
    return result

def down_sampling5(data, path):
    # Read the original Excel file

    # Assume the data has three columns named 'Column1', 'Column2', 'Column3'
    column_names = ['Force_X', 'Force_Y', 'Force_Z']  # Replace with actual column names

    # Define length of downsampled data
    desired_length = 9730

    # Downsample each column
    downsampled_data = {}
    for column in column_names:
        downsampled_data[column] = resample(data[column], desired_length)

    # Store downsampled data in DataFrame
    downsampled_df = pd.DataFrame(downsampled_data)
    return downsampled_df

# Denoising
'''
First, generate a signal containing high-frequency and low-frequency noise. 
Then use signal.butter to design a band-pass filter with cutoff frequencies to remove high- and low-frequency noise. 
Apply the filter to remove aliasing errors, and finally use a high-pass filter to remove low-frequency noise.
'''
def removeNoise(data, fs):
    from scipy import signal
    # Band-pass filter to remove high-frequency noise
    nyquist_freq = 0.5 * fs  # Nyquist frequency
    lowcut_freq = 1  # Cutoff frequency
    highcut_freq = 45
    filter_order = 5  # Filter order
    b, a = signal.butter(filter_order, [lowcut_freq / nyquist_freq, highcut_freq / nyquist_freq], 'band')
    x_filtered = signal.filtfilt(b, a, data)
    # High-pass filter to remove low-frequency noise
    cutoff_freq = 0.5  # Cutoff frequency
    b, a = signal.butter(filter_order, cutoff_freq / nyquist_freq, 'highpass')
    x_filtered = signal.filtfilt(b, a, x_filtered)
    return x_filtered

# Median filtering + baseline removal
'''
Use signal.medfilt for median filtering to remove baseline.
'''
def removeBaseLine(data, num):
    from scipy import signal
    baseline = signal.medfilt(data, kernel_size=num)  # Median filter to remove baseline
    x_nobaseline = data - baseline
    return x_nobaseline

# Normalization to (-1,1)
'''
Generate a signal, use np.abs to compute absolute values, then divide by the max value to normalize to the range [-1, 1].
'''
def normalization1(data):
    x_norm = data / np.max(np.abs(data))
    return x_norm

# Normalization to [0, 1]
def normalization2(input_list):
    data = input_list.iloc[:, :16]
    for pp in range(16):
        min_val = min(data.iloc[:, pp])
        max_val = max(data.iloc[:, pp])
        data.iloc[:, pp] = [(x - min_val) / (max_val - min_val) for x in data.iloc[:, pp]]
    return data

def normalization3(df):
    # Normalize first 8 columns by a max value of 4096
    df.iloc[:, :16] = df.iloc[:, :16] / df.iloc[:, :16].max()
    return df

# Remove slow trends
'''
Use signal.savgol_filter to apply Savitzky-Golay filtering to remove slow trends.
'''
def removeSlowChange(data):
    from scipy import signal
    trend = signal.savgol_filter(data, window_length=101, polyorder=3)  # Remove slow trends
    x_noslow = data - trend
    return x_noslow

# Get data from .xls file
def getDataXls(path, p, x):
    excel = xlrd.open_workbook(path)  # Open Excel file
    sheet = excel.sheet_by_index(0)  # Get worksheet
    data = []
    i = 1
    for col in sheet.col_values(ord(p)-ord('A')):
        if i <= x:
            i += 1
            continue
        data.append(col)
    return data

# Get data from .xlsx file
def getDataXlsx(path, str, x):
    from openpyxl import load_workbook
    wb1 = load_workbook(path)
    sheets = wb1.worksheets  # Get all sheets
    sheet1 = sheets[0]
    data = []
    i = 0
    for col in sheet1[str]:
        if i <= x:
            i += 1
            continue
        data.append(col.value)
    return data

# Read a .dat file
def readDat(path):
    with open(path, 'rb') as f:
        data = pickle.load(f)
    print(f"Unpickling:\n{data}\n")

# Save data to .xls file
def saveDataXls(path, data):
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('data', cell_overwrite_ok=True)
    for i in range(len(data)):
        sheet.write(i, 0, data[i])
    book.save(path)

# Save data to .xlsx file
def saveDataXlsx(path, data):
    data = pd.DataFrame(data)
    # Save filtered data to new Excel file
    data.to_excel(path, index=False)

# Remove outliers
def remove_outliers(group, threshold=3):
    # Calculate mean and standard deviation of the group
    mean = group.mean()
    std = group.std()
    threshold = threshold * std
    mask = (group < (mean - threshold)) | (group > (mean + threshold))
    avg_greater_than_one = np.mean(group[(group >= (mean - threshold)) & (group <= (mean + threshold))])
    group[mask] = avg_greater_than_one
    return group

def removeBig(data):
    # Divide data into groups of 10 and remove outliers in each group
    grouped_data = [remove_outliers(data[i:i+20]) for i in range(0, len(data), 20)]
    # Combine processed data into new list
    processed_data = []
    for group in grouped_data:
        processed_data.extend(group)
    return processed_data

# Fill missing data based on timestamp
def FillData(df):
    # Assume first column is the time column, with time in milliseconds
    df['timestamp'] = pd.to_numeric(df['timestamp'], errors='coerce')  # Convert to numeric
    # Get min and max time values
    min_time = df['timestamp'].min()
    max_time = df['timestamp'].max()
    # Create complete time series
    complete_time_series = pd.Series(np.arange(min_time, max_time + 10, 10))
    # Set time column as index
    df.set_index('timestamp', inplace=True)
    # Reindex to fill missing time points
    df = df.reindex(complete_time_series)
    df.index.name = 'timestamp'  # Rename index column as Time
    df.fillna('', inplace=True)  # Fill missing values with 0 or other values if no interpolation needed
    return df

# Upsampling
def upSample(df, target_rows):
    original_rows = df.shape[0]
    # Raise error if target rows < original rows
    if target_rows <= original_rows:
        raise ValueError("Target row count must be greater than original row count for upsampling.")
    # Generate new indices for interpolation
    new_index = np.linspace(0, original_rows - 1, target_rows)
    df_resampled = pd.DataFrame(index=new_index)
    for column in df.columns:
        df_resampled[column] = np.interp(new_index, np.arange(original_rows), df[column])
    df_resampled.reset_index(drop=True, inplace=True)
    return df_resampled

def upSample2(df, A, B):
    # Calculate original and target time points
    original_time = np.arange(len(df)) / A
    target_time = np.arange(len(df) * (B / A)) / B
    upsampled_data = df.apply(lambda x: interp1d(original_time, x, kind='linear', bounds_error=False, fill_value='extrapolate')(target_time))
    upsampled_df = pd.DataFrame(upsampled_data)
    return upsampled_df

# Get middle value of each segment and calculate its average (used for ground reaction force data to calculate weight "without bottom")
def get_middle_of_each_segment(data, threshold):
    data1 = data['Force_Y'].astype(float)
    data1 = data1.tolist()
    is_bottom = []
    for x in data1:
        if x < threshold:
            is_bottom.append(True)
        else:
            is_bottom.append(False)
    segments = []
    current_segment = []
    # Traverse data to identify continuous bottom segments
    for i, value in enumerate(is_bottom):
        if value:  # Current value < threshold, part of the bottom line
            current_segment.append(data.iloc[i, 3])
        elif current_segment:  # If not part of the bottom line, save segment
            if len(current_segment) > 100:
                middle_index = len(current_segment) // 2  # Get middle part
                segments.append(current_segment[middle_index])
            current_segment = []
    if current_segment:  # Process last segment if present
        middle_index = len(current_segment) // 2
        segments.append(current_segment[middle_index])
    print('segments', segments)
    return segments

# Define function to read each sheet, output column length, incremental data, and save it
# Define function to read each sheet, process data, and save it
def process_excel(file_path):
    # Create a new Excel workbook to store modified data
    output_workbook = Workbook()
    # Load input Excel file
    excel_data = pd.ExcelFile(file_path)
    # Process each worksheet
    for sheet_name in excel_data.sheet_names:
        # Read the current worksheet
        df = excel_data.parse(sheet_name)
        # Length of first 8 columns
        target_length = len(df.iloc[:, 0].dropna())
        # Downsample the 9th column to the length of the first 8 columns
        ninth_col = df.iloc[:, 8].dropna().values  # Get 9th column data
        if len(ninth_col) > target_length:
            downsampled_ninth_col = np.linspace(0, len(ninth_col) - 1, target_length).astype(int)
            ninth_col = ninth_col[downsampled_ninth_col]
        elif len(ninth_col) < target_length:
            # If length of 9th column < length of first 8 columns, pad it
            ninth_col = np.pad(ninth_col, (0, target_length - len(ninth_col)), 'constant', constant_values=np.nan)
        # Merge first 8 columns and processed 9th column into new DataFrame
        new_df = df.iloc[:target_length, :8].copy()  # Original data of first 8 columns
        new_df['Column9'] = ninth_col  # Add downsampled 9th column data
        # Append modified data to new workbook
        new_sheet = output_workbook.create_sheet(title=sheet_name)
        new_sheet.append(new_df.columns.tolist())  # Add header
        for r_idx, row in enumerate(new_df.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                new_sheet.cell(row=r_idx, column=c_idx, value=value)
    # Remove default empty worksheet
    if "Sheet" in output_workbook.sheetnames:
        output_workbook.remove(output_workbook["Sheet"])
    return output_workbook

if __name__ == '__main__':
    # Define path to folder
    folder_path = 'E:\\path\\to\\your\\folder'

    path1 = folder_path + '\\traindata.xlsx'
    path2 = folder_path + '\\traindata_preprocessing.xlsx'

    # Read data in .xls format, specify column, and skip initial rows
    # data = getDataXls(path1,'C',3)
    # Read data in .xlsx format, specify column, and skip initial rows
    # data = getDataXlsx(path1,'I',1)
    # data = pd.read_excel(path1, usecols=range(16))
    data = pd.read_excel(path1)

    # Read data in .csv format
    # data = pd.read_csv(path1)
    # data = data.drop(index=0) # Delete the second row
    # print(data)

    # Read data in .tsv format, skip specified initial rows
    # data = pd.read_csv(path1, sep='\t', skiprows = 26)
    # data = data.iloc[:,2:5]
    # print(data)


    # fs = 102.4  # Sampling rate
    # fs = 100  # Sampling rate
    # order = 5  # Order
    # cutoff1 = 5  # Cutoff frequency 1
    # cutoff2 = float(50)  # Cutoff frequency 2
    # func = butter_lowpass_filter # Filter method

    cutoff = 20  # Cutoff frequency 1
    cutoff2 = float(50)  # Cutoff frequency 2
    fs = 100  # Sampling rate
    order = 1  # Order

    # Apply low-pass filter to each column. Lower cutoff filters out high frequencies more effectively
    # data = data.apply(lambda col: filter_column(col, cutoff, fs))
    # print(data)

    # Low-pass filter
    # print(data.iloc[:,24:25].values.flatten().tolist())
    # data = butter_lowpass_filter(data.iloc[:,24:25].values.flatten().tolist(), cutoff=15, fs=100)

    # High-pass filter; retains high-frequency fluctuations, removes overall low-frequency trends
    # data = butter_highpass_filter(data,40,None,fs,2)

    # Band-pass filter
    # data = butter_bandpass_filter(data,40000,80000,fs,3)

    # Band-stop filter
    # data = butter_bandstop_filter(data,2,49,fs,5)

    # Remove high- and low-frequency noise
    # removeNoise(data,100)

    # Median filter + baseline removal; 30% of sampling rate is a suitable window width (odd number)
    # data = removeBaseLine(data,31)

    # Remove slow trend
    # data = removeSlowChange(data)

    # Normalize to (-1,1)
    # data = normalization1(data)

    # Normalize by max value of 4096
    # data = normalization3(data)

    # Normalize to [0,1] by min and max values
    # data = normalization2(data)

    # Downsample
    # data = lowerHZ34(data,100.0,99.85)
    #
    # data = normalization1(data)
    data = down_sampling4(data, original_sampling_rate = 1200, target_sampling_rate = 100)

    # Downsample each sheet
    # output_workbook = process_excel(path1)
    # output_workbook.save(path2)

    # Upsample
    # data = upSample(data,81383)
    # data = upSample2(data,40,100)

    # Fill missing data based on time window
    # data = FillData(data)


    # Get middle values of each segment and calculate their average (used to calculate "baseline-removed" weight in ground reaction force data)
    # threshold = -200 # Threshold
    # middle_values = get_middle_of_each_segment(data, threshold)
    # average_middle_values = np.mean(middle_values)
    # print('Average value:', average_middle_values)



    # Save data in .xls format
    # saveDataXls(path2,data)

    # Save data in .xlsx format
    saveDataXlsx(path2,data)

    # Save in .csv format
    # data.to_excel(path2)

